#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
沃尔玛呆滞库存明细表更新工具 v10
修复版本：
1. 使用更健壮的列名匹配逻辑
2. 增强SKU提取逻辑
3. 使用中文括号匹配
4. 原有记录全部保留
5. 只对新增记录进行SKU检验
6. 只删除新增的全零记录
"""

import openpyxl
from openpyxl.styles import Alignment, Font
import os
import re
import warnings
warnings.filterwarnings('ignore')


def find_sheet_name(sheets, keywords):
    """根据关键词查找sheet名称"""
    for sheet_name in sheets:
        if any(keyword in sheet_name for keyword in keywords):
            return sheet_name
    return None


def get_cell_value(cell):
    """获取单元格值，处理空值"""
    if cell.value is None:
        return 0
    val = str(cell.value).strip()
    if val == '' or val.lower() in ('nan', '#n/a', '#na', 'none', ''):
        return 0
    try:
        if val.startswith('='):
            return 0
        return float(val)
    except (ValueError, TypeError):
        return val


def get_numeric_value(cell):
    """获取数值"""
    if cell is None or cell.value is None:
        return 0
    val = str(cell.value).strip()
    if val == '' or val.lower() in ('nan', '#n/a', '#na', 'none', ''):
        return 0
    try:
        if val.startswith('='):
            return 0
        return float(val)
    except (ValueError, TypeError):
        return 0


def load_product_reference(product_file_path):
    """从产品资料表中加载sku集合和sku到品名的映射"""
    sku_set = set()
    sku_to_name = {}  # SKU到品名的映射

    try:
        wb_product = openpyxl.load_workbook(product_file_path, read_only=True)
        sheet_names = wb_product.sheetnames
        print(f"产品资料表sheet列表: {sheet_names}")

        ws = wb_product[sheet_names[0]]

        # 查找SKU列和品名列
        sku_col_idx = None
        name_col_idx = None
        for col_idx in range(1, 10):
            header = ws.cell(row=1, column=col_idx).value
            if header:
                header_str = str(header).strip()
                if (header_str == 'SKU' or header_str == 'sku') and sku_col_idx is None:
                    sku_col_idx = col_idx
                elif header_str in ['品名', '名称', 'name', 'Name'] and name_col_idx is None:
                    name_col_idx = col_idx

        print(f"产品资料表SKU列位置: {sku_col_idx}, 品名列位置: {name_col_idx}")

        if sku_col_idx:
            row_idx = 2
            while True:
                sku_cell = ws.cell(row=row_idx, column=sku_col_idx)
                sku_val = str(sku_cell.value).strip() if sku_cell.value else ''

                if not sku_val:
                    break

                if sku_val:
                    sku_set.add(sku_val)

                # 获取品名
                if name_col_idx and sku_val:
                    name_cell = ws.cell(row=row_idx, column=name_col_idx)
                    name_val = str(name_cell.value).strip() if name_cell.value else ''
                    if name_val:
                        sku_to_name[sku_val] = name_val

                row_idx += 1

            print(f"成功加载 {len(sku_set)} 个产品SKU，{len(sku_to_name)} 个品名映射")

        wb_product.close()

    except Exception as e:
        print(f"读取产品资料表时出错: {e}")

    return sku_set, sku_to_name


def extract_sku_smart(msku, sku_set):
    """
    更智能的sku提取逻辑
    """
    if not msku:
        return '', False

    if not sku_set:
        # 如果没有产品资料表，使用简单提取
        parts = msku.split('-')
        if len(parts) >= 2:
            return parts[1], False
        elif len(parts) >= 1:
            return parts[0], False
        return msku, False

    # 方式1：按"-"分割，逐个精确匹配
    parts = msku.split('-')

    # 去除空字符串
    parts = [p.strip() for p in parts if p.strip()]

    # 精确匹配
    for part in parts:
        if part in sku_set:
            return part, True

    # 方式2：去除引号、空格等特殊字符后再匹配
    cleaned_parts = []
    for part in parts:
        cleaned = part.replace('"', '').replace("'", '').replace(' ', '')
        if cleaned:
            cleaned_parts.append(cleaned)

    for part in cleaned_parts:
        if part in sku_set:
            return part, True

    # 方式3：模糊匹配 - 尝试子串匹配
    # 最小匹配长度为4
    for part in cleaned_parts:
        if len(part) >= 4:
            for sku in sku_set:
                if part in sku:
                    return sku, True
                if sku in part:
                    # 检查匹配度
                    match_ratio = len(part) / len(sku) if sku else 0
                    if match_ratio >= 0.5:
                        return sku, True
                        break

    # 方式4：提取看起来像sku的部分（包含字母和数字的组合）
    for part in parts:
        if re.search(r'[a-zA-Z]', part) and re.search(r'\d', part):
            cleaned = part.replace('"', '').replace("'", '').replace(' ', '')
            if cleaned in sku_set:
                return cleaned, True

    # 方式5：返回最长的部分（通常是sku）
    if cleaned_parts:
        longest = max(cleaned_parts, key=len)
        # 如果最长的部分长度 >= 4，尝试匹配
        if len(longest) >= 4:
            if longest in sku_set:
                return longest, True
            # 模糊匹配
            for sku in sku_set:
                if longest in sku or sku in longest:
                    match_ratio = len(longest) / len(sku) if sku else 0
                    if match_ratio >= 0.5:
                        return sku, True

        return longest, False

    return '', False


def update_walmart_inventory(file_path):
    """更新沃尔玛滞库存明细表"""

    print(f"正在处理文件: {file_path}")
    print("=" * 80)

    # 加载产品资料表
    sku_set = set()
    sku_to_name = {}  # SKU到品名的映射
    script_dir = os.path.dirname(os.path.abspath(__file__))
    product_file = os.path.join(script_dir, 'assets', '产品资料.xlsx')

    if os.path.exists(product_file):
        print(f"找到产品资料表: {product_file}")
        sku_set, sku_to_name = load_product_reference(product_file)
    else:
        print(f"未找到产品资料表: {product_file}")

    # 读取Excel文件
    wb = openpyxl.load_workbook(file_path)
    sheets = wb.sheetnames
    print(f"Sheet列表: {sheets}")

    # 查找各个sheet
    inventory_sheet_name = None
    sz_stock_sheet_name = None
    wfs_stock_sheet_name = None
    sales_sheet_name = None
    po_sheet_name = None

    # 库存明细表 - 第2个sheet
    inventory_sheet_name = sheets[1] if len(sheets) > 1 else None

    # 深圳仓库存 - 模糊匹配
    sz_stock_sheet_name = find_sheet_name(sheets, ['深圳仓', '深圳', '仓库'])

    # WFS库存 - 模糊匹配
    wfs_stock_sheet_name = find_sheet_name(sheets, ['WFS库存', 'WFS'])

    # 销量明细 - 第5个sheet（索引4）
    sales_sheet_name = sheets[4] if len(sheets) > 4 else None

    # 采购订单在途 - 模糊匹配
    po_sheet_name = find_sheet_name(sheets, ['采购订单', '采购', '在途'])

    print(f"\n识别到的sheet:")
    print(f"  库存明细表: {inventory_sheet_name}")
    print(f"  深圳仓库存: {sz_stock_sheet_name}")
    print(f"  WFS库存: {wfs_stock_sheet_name}")
    print(f"  销量明细: {sales_sheet_name}")
    print(f"   采购订单在途: {po_sheet_name}")

    # ========== 第0步：记录原有记录的店铺+msku组合 ==========
    print("\n" + "=" * 80)
    print("第0步：记录原有记录的店铺+msku组合")
    print("=" * 80)

    inventory_sheet = wb[inventory_sheet_name]

    existing_keys = set()
    original_row_count = 0

    row_idx = 3
    while True:
        store_cell = inventory_sheet.cell(row=row_idx, column=1)
        msku_cell = inventory_sheet.cell(row=row_idx, column=2)

        store = str(store_cell.value).strip() if store_cell.value else ''
        msku = str(msku_cell.value).strip() if msku_cell.value else ''

        if not store and not msku:
            break

        if store and msku:
            existing_keys.add(f"{store}{msku}")
            original_row_count += 1

        row_idx += 1

    print(f"原有记录数: {original_row_count}")
    print(f"原有店铺+msku组合数: {len(existing_keys)}")
    print(f"最后一行原有记录: {original_row_count + 2}")

    # ========== 第1步：准备WFS库存数据 ==========
    print("\n" + "=" * 80)
    print("第1步：准备WFS库存数据")
    print("=" * 80)

    wfs_sheet = wb[wfs_stock_sheet_name]
    wfs_col_map = {}

    # 查找WFS库存表的关键列（第1行是表头）
    for col_idx in range(1, 40):
        cell = wfs_sheet.cell(row=1, column=col_idx)
        if cell.value:
            col_name = str(cell.value).strip()
            if '仓库' in col_name:
                wfs_col_map['仓库'] = col_idx
            elif 'msku' in col_name:
                wfs_col_map['msku'] = col_idx
            elif col_name == 'GTIN码':
                wfs_col_map['GTIN码'] = col_idx
            elif '平台' in col_name and '商品' in col_name and 'ID' in col_name:
                wfs_col_map['平台商品ID'] = col_idx
            elif '品名' in col_name and 'ID' not in col_name:
                wfs_col_map['品名'] = col_idx
            elif col_name == 'sku':
                wfs_col_map['sku'] = col_idx
            elif '商品状态' in col_name and '库龄' not in col_name:
                wfs_col_map['商品状态'] = col_idx
            elif 'WFS可售(新)' in col_name and '数量' in col_name:
                wfs_col_map['WFS可售(新)(数量)'] = col_idx
            elif '无法入库' in col_name and '数量' in col_name:
                wfs_col_map['无法入库(数量)'] = col_idx
            elif '标发在途' in col_name and '数量' in col_name:
                wfs_col_map['标发在途(数量)'] = col_idx
            elif '3个月内库龄' in col_name and '数量' in col_name:
                wfs_col_map['3个月内库龄(数量)'] = col_idx
            elif '3-6个月库龄' in col_name and '数量' in col_name:
                wfs_col_map['3-6个月库龄(数量)'] = col_idx
            elif '6个月以上库龄' in col_name and '数量' in col_name:
                wfs_col_map['6个月以上库龄(数量)'] = col_idx
            elif '12个月以上库龄' in col_name and '数量' in col_name:
                wfs_col_map['12个月以上库龄(数量)'] = col_idx

    print(f"WFS库存列映射: {wfs_col_map}")

    # 创建WFS库存映射
    wfs_dict = {}
    row_idx = 2
    while True:
        warehouse_cell = wfs_sheet.cell(row=row_idx, column=wfs_col_map.get('仓库', 1))
        msku_cell = wfs_sheet.cell(row=row_idx, column=wfs_col_map.get('msku', 2))

        warehouse = str(warehouse_cell.value).strip() if warehouse_cell.value else ''
        msku = str(msku_cell.value).strip() if msku_cell.value else ''

        if not warehouse and not msku:
            break

        if warehouse and msku:
            # 直接拼接，不使用下划线
            key = f"{warehouse}{msku}"
            # 获取GTIN码时确保是文本格式
            gtin_cell = wfs_sheet.cell(row=row_idx, column=wfs_col_map.get('GTIN码', 10))
            gtin_value = str(gtin_cell.value).strip() if gtin_cell.value and gtin_cell.value != '' else ''

            wfs_dict[key] = {
                '仓库': warehouse,
                'GTIN码': gtin_value,
                '平台商品ID': get_cell_value(wfs_sheet.cell(row=row_idx, column=wfs_col_map.get('平台商品ID', 11))),
                '品名': get_cell_value(wfs_sheet.cell(row=row_idx, column=wfs_col_map.get('品名', 12))),
                'sku': get_cell_value(wfs_sheet.cell(row=row_idx, column=wfs_col_map.get('sku', 13))),
                '商品状态': get_cell_value(wfs_sheet.cell(row=row_idx, column=wfs_col_map.get('商品状态', 14))),
                'WFS可售(新)(数量)': get_numeric_value(wfs_sheet.cell(row=row_idx, column=wfs_col_map.get('WFS可售(新)(数量)', 18))),
                '无法入库(数量)': get_numeric_value(wfs_sheet.cell(row=row_idx, column=wfs_col_map.get('无法入库(数量)', 19))),
                '标发在途(数量)': get_numeric_value(wfs_sheet.cell(row=row_idx, column=wfs_col_map.get('标发在途(数量)', 20))),
                '3个月内库龄(数量)': get_numeric_value(wfs_sheet.cell(row=row_idx, column=wfs_col_map.get('3个月内库龄(数量)', 4))),
                '3-6个月库龄(数量)': get_numeric_value(wfs_sheet.cell(row=row_idx, column=wfs_col_map.get('3-6个月库龄(数量)', 5))),
                '6个月以上库龄(数量)': get_numeric_value(wfs_sheet.cell(row_idx, column=wfs_col_map.get('6个月以上库龄(数量)', 6))),
                '12个月以上库龄(数量)': get_numeric_value(wfs_sheet.cell(row=row_idx, column=wfs_col_map.get('12个月以上库龄(数量)', 7))),
            }

        row_idx += 1

    print(f"WFS库存记录数: {len(wfs_dict)}")

    # ========== 第2步：准备销量明细数据 ==========
    print("\n" + "=" * 80)
    print("第2步：准备销量明细数据")
    print("=" * 80)

    sales_sheet = wb[sales_sheet_name]
    sales_col_map = {}

    # 查找销量明细表的关键列（第1行是表头）
    for col_idx in range(1, 50):
        cell = sales_sheet.cell(row=1, column=col_idx)
        if cell.value:
            col_name = str(cell.value).strip()
            if col_name == 'MSKU':
                sales_col_map['MSKU'] = col_idx
            elif col_name == 'SKU':
                sales_col_map['SKU'] = col_idx
            elif col_name == '品名':
                sales_col_map['品名'] = col_idx
            elif col_name == '店铺':
                sales_col_map['店铺'] = col_idx
            elif col_name == '小计':
                sales_col_map['小计'] = col_idx

    print(f"销量明细列映射: {sales_col_map}")

    # 创建销量明细映射
    sales_dict = {}
    row_idx = 2
    while True:
        msku_cell = sales_sheet.cell(row=row_idx, column=sales_col_map.get('MSKU', 4))
        sku_cell = sales_sheet.cell(row=row_idx, column=sales_col_map.get('SKU', 6))
        store_cell = sales_sheet.cell(row=row_idx, column=sales_col_map.get('店铺', 3))

        msku = str(msku_cell.value).strip() if msku_cell.value else ''
        store = str(store_cell.value).strip() if store_cell.value else ''

        if not msku:
            break

        # 从"店铺"列提取店铺名称
        if not store:
            if '-' in msku:
                store = msku.split('-')[0]
            else:
                store = msku

        # 获取sku，如果为空则从msku中提取
        sku = str(sku_cell.value).strip() if sku_cell.value and sku_cell.value != '' else ''
        if not sku:
            sku, _ = extract_sku_smart(msku, sku_set)

        if store and msku:
            # 直接拼接，不使用下划线
            key = f"{store}{msku}"
            sales_dict[key] = {
                '店铺': store,
                '销量': get_numeric_value(sales_sheet.cell(row=row_idx, column=sales_col_map.get('小计', 13))),
                'SKU': sku,
                '品名': get_cell_value(sales_sheet.cell(row=row_idx, column=sales_col_map.get('品名', 7))),
                'msku': msku,
            }

        row_idx += 1

    print(f"销量明细记录数: {len(sales_dict)}")

    # ========== 第3步：准备深圳仓库存数据 ==========
    print("\n" + "=" * 80)
    print("第3步：准备深圳仓库存数据")
    print("=" * 80)

    sz_sheet = wb[sz_stock_sheet_name]
    sz_col_map = {}

    # 查找深圳仓库存表的关键列（第1行是表头）
    for col_idx in range(1, 15):
        cell = sz_sheet.cell(row=1, column=col_idx)
        if cell.value:
            col_name = str(cell.value).strip()
            if col_name == 'SKU':
                sz_col_map['SKU'] = col_idx
            elif col_name == '实际可用':
                sz_col_map['实际可用'] = col_idx

    print(f"深圳仓库存列映射: {sz_col_map}")

    # 创建深圳仓库存映射
    sz_stock_dict = {}
    row_idx = 2
    while True:
        sku_cell = sz_sheet.cell(row=row_idx, column=sz_col_map.get('SKU', 1))
        sku = str(sku_cell.value).strip() if sku_cell.value else ''

        if not sku:
            break

        sz_stock_dict[sku] = get_numeric_value(sz_sheet.cell(row=row_idx, column=sz_col_map.get('实际可用', 10)))
        row_idx += 1

    print(f"深圳仓库存记录数: {len(sz_stock_dict)}")

    # ========== 第4步：准备采购订单在途数据 ==========
    print("\n" + "=" * 80)
    print("第4步：准备采购订单在途数据")
    print("=" * 80)

    po_sheet = wb[po_sheet_name]
    po_col_map = {}

    # 查找采购订单在途表的关键列（第1行是表头）
    for col_idx in range(1, 35):
        cell = po_sheet.cell(row=1, column=col_idx)
        if cell.value:
            col_name = str(cell.value).strip()
            if col_name == 'SKU':
                po_col_map['SKU'] = col_idx
            elif col_name == '未入库量':
                po_col_map['未入库量'] = col_idx

    print(f"采购订单在途列映射: {po_col_map}")

    # 按SKU汇总采购订单在途量
    po_dict = {}
    row_idx = 2
    while True:
        sku_cell = po_sheet.cell(row=row_idx, column=po_col_map.get('SKU', 7))
        sku = str(sku_cell.value).strip() if sku_cell.value else ''

        if not sku:
            break

        qty = get_numeric_value(po_sheet.cell(row=row_idx, column=po_col_map.get('未入库量', 19)))
        if sku in po_dict:
            po_dict[sku] += qty
        else:
            po_dict[sku] = qty
        row_idx += 1

    print(f"采购订单在途记录数: {len(po_dict)}")

    # ========== 第5步：处理库存明细表数据 ==========
    print("\n" + "=" * 80)
    print("第5步：处理库存明细表数据")
    print("=" * 80)

    inventory_sheet = wb[inventory_sheet_name]
    inventory_col_map = {}

    # 查找库存明细表的关键列（第2行是表头）
    for col_idx in range(1, 100):
        cell = inventory_sheet.cell(row=2, column=col_idx)
        if cell.value:
            col_name = str(cell.value).strip()
            if col_name == '店铺' and 'MSKU' not in col_name:
                inventory_col_map['店铺'] = col_idx
            elif col_name == 'msku' and '店铺' not in col_name:
                inventory_col_map['msku'] = col_idx
            elif col_name == '店铺&MSKU':
                inventory_col_map['店铺&MSKU'] = col_idx
            elif col_name == 'GTIN码':
                inventory_col_map['GTIN码'] = col_idx
            elif '平台' in col_name and '商品' in col_name and 'ID' in col_name:
                inventory_col_map['平台商品ID'] = col_idx
            elif col_name == '品名' and 'ID' not in col_name:
                inventory_col_map['品名'] = col_idx
            elif col_name == '运营':
                inventory_col_map['运营'] = col_idx
            elif col_name == 'sku':
                inventory_col_map['sku'] = col_idx
            elif col_name == '商品状态':
                inventory_col_map['商品状态'] = col_idx
            elif 'WFS可售(新)' in col_name and '数量' in col_name:
                inventory_col_map['WFS可售(新)(数量)'] = col_idx
            elif '无法入库' in col_name and '数量' in col_name:
                inventory_col_map['无法入库(数量)'] = col_idx
            elif '标发在途' in col_name and '数量' in col_name:
                inventory_col_map['标发在途(数量)'] = col_idx
            elif '深圳仓库存' in col_name:
                inventory_col_map['深圳仓库存'] = col_idx
            elif '采购订单在途' in col_name:
                inventory_col_map['采购订单在途'] = col_idx
            elif '总库存' in col_name:
                inventory_col_map['总库存'] = col_idx
            elif 'WFS在库周转' in col_name:
                inventory_col_map['WFS在库周转'] = col_idx
            elif 'WFS在途+在库周转' in col_name:
                inventory_col_map['WFS在途+在库周转'] = col_idx
            elif '总周转天数（不含采购订单）' in col_name:
                inventory_col_map['总周转天数（不含采购订单）'] = col_idx
            elif '3个月内库龄' in col_name and '数量' in col_name:
                inventory_col_map['3个月内库龄(数量)'] = col_idx
            elif '3-6个月库龄' in col_name and '数量' in col_name:
                inventory_col_map['3-6个月库龄(数量)'] = col_idx
            elif '6个月以上库龄' in col_name and '数量' in col_name:
                inventory_col_map['6个月以上库龄(数量)'] = col_idx
            elif '12个月以上库龄' in col_name and '数量' in col_name:
                inventory_col_map['12个月以上库龄(数量)'] = col_idx

    print(f"库存明细表列映射: {len(inventory_col_map)}")

    # 检查是否有销量列，没有则添加
    sales_column_idx = None
    for col_idx in range(1, 200):
        cell = inventory_sheet.cell(row=2, column=col_idx)
        if cell.value and str(cell.value).strip() == sales_sheet_name:
            sales_column_idx = col_idx
            break

    if sales_column_idx is None:
        # 添加新的销量列（在最后）
        sales_column_idx = inventory_sheet.max_column + 1
        # 设置列名
        inventory_sheet.cell(row=2, column=sales_column_idx, value=sales_sheet_name)
        inventory_sheet.cell(row=2, column=sales_column_idx).alignment = Alignment(horizontal='center')
        print(f"添加了新的销量列: {sales_sheet_name} (列 {sales_column_idx})")

    print(f"销量列位置: {sales_column_idx}")

    # 定义标红/绿样式
    red_font = Font(color='FF0000', bold=True)
    green_font = Font(color='00FF00', bold=True)

    # 更新库存明细表中的现有记录（不删除）
    print(f"\n更新现有记录（保留原有 {original_row_count} 条）...")
    row_idx = 3
    while row_idx <= original_row_count + 2:  # 只处理原有记录
        store_cell = inventory_sheet.cell(row=row_idx, column=1)
        msku_cell = inventory_sheet.cell(row=row_idx, column=2)

        store = str(store_cell.value).strip() if store_cell.value else ''
        msku = str(msku_cell.value).strip() if msku_cell.value else ''

        if not store and not msku:
            break

        sku_cell = inventory_sheet.cell(row=row_idx, column=8)
        sku = str(sku_cell.value).strip() if sku_cell.value else ''

        # 查找店铺+msku的键
        key1 = f"{store}{msku}"
        key = None
        if key1 in wfs_dict or key1 in sales_dict:
            key = key1

        # 如果库存明细表中没有店铺，尝试用sku来匹配
        if not key and sku:
            for wfs_key in wfs_dict.keys():
                if sku == str(wfs_dict[wfs_key].get('sku', '')):
                    # 从wfs_key中提取店铺部分
                    store_part = wfs_dict[wfs_key].get('仓库', '')
                    inventory_sheet.cell(row=row_idx, column=1, value=store_part)
                    if '店铺&MSKU' in inventory_col_map:
                        inventory_sheet.cell(row=row_idx, column=inventory_col_map['店铺&MSKU'], value=f"{store_part}{msku}")
                    key = wfs_key
                    store = store_part
                    break

        # 从WFS库存更新数据（优先）
        if key and key in wfs_dict:
            wfs_data = wfs_dict[key]
            for field, col_idx in inventory_col_map.items():
                if field in wfs_data and wfs_data.get(field, '') != 0 and wfs_data.get(field, '') != '':
                    inventory_sheet.cell(row=row_idx, column=col_idx, value=wfs_data[field])

        # 从销量明细更新销量
        if key and key in sales_dict:
            sales_data = sales_dict[key]
            inventory_sheet.cell(row=row_idx, column=sales_column_idx, value=sales_data['销量'])

        # 从深圳仓库存更新（以sku为匹配）
        if sku and sku in sz_stock_dict:
            inventory_sheet.cell(row=row_idx, column=inventory_col_map['深圳仓库存'], value=sz_stock_dict[sku])

        # 从采购订单在途更新（以sku为匹配）
        if sku and sku in po_dict:
            inventory_sheet.cell(row=row_idx, column=inventory_col_map['采购订单在途'], value=po_dict[sku])

        row_idx += 1

    print(f"更新完成，原有记录共 {original_row_count} 条（全部保留）")

    # ========== 第6步：添加新记录 ==========
    print("\n" + "=" * 80)
    print("第6步：添加新记录")
    print("=" * 80)

    # 记录所有出现过的店铺+msku组合
    all_store_msku_keys = set()
    for key in wfs_dict.keys():
        all_store_msku_keys.add(key)
    for key in sales_dict.keys():
        all_store_msku_keys.add(key)

    print(f"发现 {len(all_store_msku_keys)} 个店铺+msku组合")

    # 只添加不存在的key（新增记录）
    new_rows = []
    new_row_keys = set()  # 记录新增记录的key

    for key in all_store_msku_keys:
        if key not in existing_keys:
            new_row_keys.add(key)

            # 优先从WFS库存中获取店铺名称
            store = ''
            if key in wfs_dict:
                store = wfs_dict[key].get('仓库', '')
            else:
                if key in sales_dict:
                    store = sales_dict[key].get('店铺', '')
                else:
                    if '-' in key:
                        parts = key.split('-')
                        store = parts[0]
                        for part in parts[1:]:
                            if part in sku_set:
                                store += '-' + part
                                break
                    else:
                        store = key

            # msku是key减去店铺部分
            msku = key[len(store):] if len(key) > len(store) else key

            # 按照skill逻辑提取sku（对新增记录进行SKU检验）
            sku = ''
            sku_found = False
            if sku_set:
                sku, sku_found = extract_sku_smart(msku, sku_set)

            new_row = {
                '店铺': store,
                'msku': msku,
                '店铺&MSKU': key,
                'GTIN码': '',
                '平台商品ID': '',
                '品名': '',
                '运营': '',
                'sku': sku,
                '商品状态': '',
                'WFS可售(新)(数量)': 0,
                '无法入库(数量)': 0,
                '标发在途(数量)': 0,
                '深圳仓库存': 0,
                '采购订单在途': 0,
                '总库存': 0,
                'WFS在库周转': '',
                'WFS在途+在库周转': '',
                '总周转天数（不含采购订单）': '',
                '3个月内库龄(数量)': 0,
                '3-6个月库龄(数量)': 0,
                '6个月以上库龄(数量)': 0,
                '12个月以上库龄(数量)': 0,
                sales_sheet_name: 0,
                '_sku_found': sku_found,
                '_key': key,
            }

            # 从WFS库存补充数据（优先）
            if key in wfs_dict:
                wfs_data = wfs_dict[key]
                for field in wfs_data:
                    if field in new_row:
                        # 跳过空的SKU字段，保留智能提取的SKU
                        if field == 'sku' and (not wfs_data[field] or wfs_data[field] == '' or wfs_data[field] == 0):
                            continue
                        new_row[field] = wfs_data[field]

            # 从销量明细补充数据（如果WFS中没有）
            if key in sales_dict:
                sales_data = sales_dict[key]
                new_row[sales_sheet_name] = sales_data['销量']
                # 只有当WFS中没有对应数据时，才使用销量明细的品名和sku
                if key not in wfs_dict:
                    if not new_row.get('品名') and sales_data.get('品名'):
                        new_row['品名'] = sales_data['品名']
                    # 只有当销量明细的SKU不为空，且当前SKU为空时，才使用销量明细的SKU
                    if sales_data.get('SKU') and not new_row.get('sku'):
                        new_row['sku'] = sales_data['SKU']

            # 如果有SKU且有产品资料映射，补充品名
            if new_row.get('sku') and new_row['sku'] in sku_to_name:
                if not new_row.get('品名') or new_row.get('品名') == '':
                    new_row['品名'] = sku_to_name[new_row['sku']]

            new_rows.append(new_row)

    print(f"需要新增 {len(new_rows)} 条记录")

    # 添加新行到文件
    if new_rows:
        for new_row in new_rows:
            row_idx = inventory_sheet.max_row + 1
            for field, value in new_row.items():
                if field.startswith('_'):
                    continue
                if field in inventory_col_map:
                    inventory_sheet.cell(row=row_idx, column=inventory_col_map[field], value=value)
                elif field == sales_sheet_name:
                    inventory_sheet.cell(row=row_idx, column=sales_column_idx, value=value)

    print(f"已添加 {len(new_rows)} 条新记录到文件末尾")

    # ========== 第7步：计算公式字段 ==========
    print("\n" + "=" * 80)
    print("第7步：计算公式字段")
    print("=" * 80)

    # 重新遍历所有记录计算公式
    row_idx = 3
    while row_idx <= inventory_sheet.max_row:
        store = str(inventory_sheet.cell(row=row_idx, column=1).value).strip() if inventory_sheet.cell(row=row_idx, column=1).value else ''
        msku = str(inventory_sheet.cell(row=row_idx, column=2).value).strip() if inventory_sheet.cell(row=row_idx, column=2).value else ''

        if not store and not msku:
            break

        # 获取各项数据
        wfs_qty = get_numeric_value(inventory_sheet.cell(row=row_idx, column=inventory_col_map['WFS可售(新)(数量)']))
        unable_qty = get_numeric_value(inventory_sheet.cell(row=row_idx, column=inventory_col_map['无法入库(数量)']))
        transit_qty = get_numeric_value(inventory_sheet.cell(row=row_idx, column=inventory_col_map['标发在途(数量)']))
        sz_qty = get_numeric_value(inventory_sheet.cell(row=row_idx, column=inventory_col_map['深圳仓库存']))
        po_qty = get_numeric_value(inventory_sheet.cell(row=row_idx, column=inventory_col_map['采购订单在途']))
        sales = get_numeric_value(inventory_sheet.cell(row=row_idx, column=sales_column_idx))

        # 计算总库存
        total_stock = wfs_qty + unable_qty + transit_qty + sz_qty
        inventory_sheet.cell(row=row_idx, column=inventory_col_map['总库存'], value=total_stock)

        # 计算周转天数
        if sales > 0:
            wfs_turnover = round(wfs_qty / sales * 30, 2)
            wfs_transit_turnover = round((wfs_qty + transit_qty) / sales * 30, 2)
            total_turnover_no_po = round((wfs_qty + transit_qty + sz_qty) / sales * 30, 2)
        else:
            wfs_turnover = ''
            wfs_transit_turnover = ''
            total_turnover_no_po = ''

        inventory_sheet.cell(row=row_idx, column=inventory_col_map['WFS在库周转'], value=wfs_turnover)
        inventory_sheet.cell(row=row_idx, column=inventory_col_map['WFS在途+在库周转'], value=wfs_transit_turnover)
        inventory_sheet.cell(row=row_idx, column=inventory_col_map['总周转天数（不含采购订单）'], value=total_turnover_no_po)

        row_idx += 1

    print("公式字段计算完成")

    # ========== 第8步：删除新增的全零记录 ==========
    print("\n" + "=" * 80)
    print("第8步：删除新增的全零记录")
    print("=" * 80)

    new_zero_rows_to_delete = []

    for new_row in new_rows:
        key = new_row.get('_key')
        row_idx = None

        # 查找该新增记录的行号（在原有记录之后）
        for i in range(original_row_count + 3, inventory_sheet.max_row + 1):
            store = str(inventory_sheet.cell(row=i, column=1).value).strip() if inventory_sheet.cell(row=i, column=1).value else ''
            msku = str(inventory_sheet.cell(row=i, column=2).value).strip() if inventory_sheet.cell(row=i, column=2).value else ''
            if f"{store}{msku}" == key:
                row_idx = i
                break

        if row_idx:
            # 检查是否全零
            wfs_qty = get_numeric_value(inventory_sheet.cell(row=row_idx, column=inventory_col_map['WFS可售(新)(数量)']))
            unable_qty = get_numeric_value(inventory_sheet.cell(row=row_idx, column=inventory_col_map['无法入库(数量)']))
            transit_qty = get_numeric_value(inventory_sheet.cell(row=row_idx, column=inventory_col_map['标发在途(数量)']))
            sz_qty = get_numeric_value(inventory_sheet.cell(row=row_idx, column=inventory_col_map['深圳仓库存']))
            po_qty = get_numeric_value(inventory_sheet.cell(row=row_idx, column=inventory_col_map['采购订单在途']))
            sales = get_numeric_value(inventory_sheet.cell(row=row_idx, column=sales_column_idx))

            is_all_zero = True
            for val in [wfs_qty, unable_qty, transit_qty, sz_qty, po_qty, sales]:
                if val not in [0, None, '0', '', 0.0]:
                    is_all_zero = False
                    break

            if is_all_zero:
                new_zero_rows_to_delete.append(row_idx)

    # 从后往前删除新增的全零记录
    for row_to_delete in sorted(new_zero_rows_to_delete, reverse=True):
        inventory_sheet.delete_rows(row_to_delete, 1)

    print(f"删除了 {len(new_zero_rows_to_delete)} 条新增的全零记录")

    # ========== 第8.5步：删除空行 ==========
    print("\n" + "=" * 80)
    print("第8.5步：删除空行")
    print("=" * 80)

    # 找出原有记录之后的空行
    empty_rows_to_delete = []
    for i in range(original_row_count + 3, inventory_sheet.max_row + 1):
        store = str(inventory_sheet.cell(i, 1).value).strip() if inventory_sheet.cell(i, 1).value else ''
        msku = str(inventory_sheet.cell(i, 2).value).strip() if inventory_sheet.cell(i, 2).value else ''
        if not store and not msku:
            empty_rows_to_delete.append(i)

    # 从后往前删除空行
    for row_to_delete in sorted(empty_rows_to_delete, reverse=True):
        inventory_sheet.delete_rows(row_to_delete, 1)

    print(f"删除了 {len(empty_rows_to_delete)} 行空行")

    # ========== 第9步：保存结果 ==========
    print("\n" + "=" * 80)
    print("第9步：保存结果")
    print("=" * 80)

    wb.save(file_path)
    print(f"\n文件已保存: {file_path}")

    # 显示更新统计
    final_total_count = inventory_sheet.max_row - 1
    added_count = final_total_count - original_row_count

    print("\n" + "=" * 80)
    print("=== 更新统计 ===")
    print("=" * 80)
    print(f"原有记录数: {original_row_count}")
    print(f"新增记录数: {len(new_rows)}")
    print(f"删除新增全零记录数: {len(new_zero_rows_to_delete)}")
    print(f"实际新增记录数: {added_count}")
    print(f"最终总记录数: {final_total_count}")
    print(f"WFS库存记录数: {len(wfs_dict)}")
    print(f"销量明细记录数: {len(sales_dict)}")
    print(f"深圳仓库存SKU数: {len(sz_stock_dict)}")
    print(f"采购订单在途SKU数: {len(po_dict)}")
    print(f"销量列名: {sales_sheet_name}")
    print("=" * 80)

    return wb


if __name__ == '__main__':
    import sys

    file_path = None

    if len(sys.argv) >= 2:
        file_path = sys.argv[1]
    else:
        # 使用默认路径
        file_path = r'C:\Users\13676\Downloads\沃尔玛呆滞库存明细表0301 (1).xlsx'

    update_walmart_inventory(file_path)
