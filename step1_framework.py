import streamlit as st
import openpyxl
from openpyxl.styles import Alignment, Font
import io
import re
import warnings

warnings.filterwarnings('ignore')

# ==========================================
# 核心逻辑函数 (V14 容错探测 + 4列过滤版)
# ==========================================

def clean_header(val):
    """清理表头：统一括号，去除空格"""
    if not val: return ""
    s = str(val).strip()
    s = s.replace('（', '(').replace('）', ')')
    s = s.replace(' ', '')
    return s

def find_exact_col(sheet, exact_name):
    """精确匹配列名（忽略中英文括号差异和空格）"""
    target = clean_header(exact_name)
    for col in range(1, 100):
        val = sheet.cell(row=1, column=col).value # WFS等表头在第1行
        if not val: 
            # 检查第二行（针对主表明细）
            val = sheet.cell(row=2, column=col).value
            if not val: continue
        
        if target in clean_header(val):
            return col
    return None

def get_real_max_row(sheet, start_row=3):
    """寻找真实的最后一行，容忍中间有空行，但连续50行空则判定为到底"""
    real_max = start_row - 1
    empty_count = 0
    for r in range(start_row, sheet.max_row + 1):
        v1 = str(sheet.cell(row=r, column=1).value or '').strip()
        v2 = str(sheet.cell(row=r, column=2).value or '').strip()
        
        if not v1 and not v2:
            empty_count += 1
            if empty_count > 50:  # 连续50行无 店铺 和 MSKU，说明真到底了
                break
        else:
            real_max = r
            empty_count = 0 # 计数器清零
            
    return real_max

def get_numeric_value(cell):
    """获取数值"""
    if cell is None or cell.value is None: return 0
    val = str(cell.value).strip()
    if val == '' or val.lower() in ('nan', '#n/a', 'none'): return 0
    try:
        if val.startswith('='): return 0
        return float(val.replace(',', ''))
    except: return 0

def load_product_reference_from_obj(product_file_obj):
    sku_set = set()
    sku_to_name = {}
    if product_file_obj is None: return sku_set, sku_to_name

    try:
        wb = openpyxl.load_workbook(product_file_obj, read_only=True)
        ws = wb[wb.sheetnames[0]]
        
        sku_col, name_col = None, None
        for col in range(1, 20):
            val = str(ws.cell(1, col).value or '').lower()
            if 'sku' in val: sku_col = col
            if val in ['品名', '名称', 'name', '品名英']: name_col = col

        if sku_col:
            empty_count = 0
            for r in range(2, ws.max_row + 1):
                sku_val = str(ws.cell(r, sku_col).value or '').strip()
                if not sku_val:
                    empty_count += 1
                    if empty_count > 50: break
                    continue
                else:
                    empty_count = 0
                
                sku_set.add(sku_val)
                if name_col:
                    name_val = str(ws.cell(r, name_col).value or '').strip()
                    if name_val: sku_to_name[sku_val] = name_val
        wb.close()
        return sku_set, sku_to_name
    except Exception as e:
        return set(), {}

def extract_sku_smart(msku, sku_set):
    """智能SKU提取"""
    if not msku: return '', False
    if not sku_set:
        parts = msku.split('-')
        return (parts[1] if len(parts)>=2 else parts[0]), False

    parts = msku.split('-')
    parts = [p.strip() for p in parts if p.strip()]

    # 精确匹配
    for p in parts:
        if p in sku_set: return p, True
    
    # 去特殊字符匹配
    cleaned = [p.replace('"', '').replace("'", '').replace(' ', '') for p in parts]
    for p in cleaned:
        if p in sku_set: return p, True

    # 模糊匹配
    for p in cleaned:
        if len(p) >= 4 and re.search(r'\d', p) and re.search(r'[a-zA-Z]', p):
            if p in sku_set: return p, True
            for sku in sku_set:
                if p in sku or sku in p:
                    if len(sku) > 0 and len(p)/len(sku) >= 0.6:
                        return sku, True
    return '', False

def process_inventory(inventory_file, product_file):
    # 1. 加载资料
    sku_set, sku_to_name = load_product_reference_from_obj(product_file)

    # 2. 加载主文件
    wb = openpyxl.load_workbook(inventory_file)
    sheets = wb.sheetnames
    
    def find_sheet(keywords):
        for s in sheets:
            if any(k in s for k in keywords): return s
        return None

    inventory_sheet_name = sheets[1] if len(sheets) > 1 else None
    sz_stock_sheet_name = find_sheet(['深圳仓', '深圳'])
    wfs_stock_sheet_name = find_sheet(['WFS库存', 'WFS'])
    sales_sheet_name = sheets[4] if len(sheets) > 4 else None
    po_sheet_name = find_sheet(['采购订单', '采购', '在途'])

    if not all([inventory_sheet_name, wfs_stock_sheet_name, sales_sheet_name]):
        st.error("❌ 找不到关键Sheet (库存明细、WFS、销量明细)，请检查文件。")
        return None

    # === 第0步：保护原有记录 ===
    inv_sheet = wb[inventory_sheet_name]
    # 使用探测器寻找真实最后一行
    original_max_row = get_real_max_row(inv_sheet, start_row=3)
    
    existing_keys = set()
    for r in range(3, original_max_row + 1):
        s = str(inv_sheet.cell(r, 1).value or '').strip()
        m = str(inv_sheet.cell(r, 2).value or '').strip()
        if s or m: existing_keys.add(f"{s}{m}")

    st.write(f"🛡️ 历史数据范围锁定：第 3 行 至 第 {original_max_row} 行")

    # === 第1步：WFS 库存 ===
    wfs_sheet = wb[wfs_stock_sheet_name]
    wfs_dict = {}
    
    w_wh = find_exact_col(wfs_sheet, '仓库') or 1
    w_msku = find_exact_col(wfs_sheet, 'msku') or 2
    w_gtin = find_exact_col(wfs_sheet, 'GTIN码')
    w_sku = find_exact_col(wfs_sheet, 'sku')
    w_name = find_exact_col(wfs_sheet, '品名')
    w_status = find_exact_col(wfs_sheet, '商品状态')
    
    w_avail = find_exact_col(wfs_sheet, 'WFS可售(新)(数量)')
    w_unable = find_exact_col(wfs_sheet, '无法入库(数量)')
    w_transit = find_exact_col(wfs_sheet, '标发在途(数量)')

    empty_count = 0
    for r in range(2, wfs_sheet.max_row + 1):
        wh = str(wfs_sheet.cell(r, w_wh).value or '').strip()
        msku = str(wfs_sheet.cell(r, w_msku).value or '').strip()
        
        # 容忍空行，不直接break
        if not wh and not msku: 
            empty_count += 1
            if empty_count > 50: break
            continue
        else:
            empty_count = 0
        
        if wh and msku:
            key = f"{wh}{msku}"
            wfs_dict[key] = {
                '仓库': wh, 'msku': msku,
                'GTIN码': str(wfs_sheet.cell(r, w_gtin).value or '') if w_gtin else '',
                'sku': str(wfs_sheet.cell(r, w_sku).value or '').strip() if w_sku else '',
                '品名': wfs_sheet.cell(r, w_name).value if w_name else '',
                '商品状态': wfs_sheet.cell(r, w_status).value if w_status else '',
                'WFS可售': get_numeric_value(wfs_sheet.cell(r, w_avail)) if w_avail else 0,
                '无法入库': get_numeric_value(wfs_sheet.cell(r, w_unable)) if w_unable else 0,
                '标发在途': get_numeric_value(wfs_sheet.cell(r, w_transit)) if w_transit else 0
            }

    # === 第2步：销量明细 ===
    sales_sheet = wb[sales_sheet_name]
    sales_dict = {}
    s_msku = find_exact_col(sales_sheet, 'MSKU') or 4
    s_store = find_exact_col(sales_sheet, '店铺') or 3
    s_subtotal = find_exact_col(sales_sheet, '小计') or 13
    s_sku = find_exact_col(sales_sheet, 'SKU')
    s_name = find_exact_col(sales_sheet, '品名')

    empty_count = 0
    for r in range(2, sales_sheet.max_row + 1):
        msku = str(sales_sheet.cell(r, s_msku).value or '').strip()
        store = str(sales_sheet.cell(r, s_store).value or '').strip()
        
        if not msku and not store:
            empty_count += 1
            if empty_count > 50: break
            continue
        else:
            empty_count = 0
            
        if not store and '-' in msku: store = msku.split('-')[0]
             
        if msku:
            key = f"{store}{msku}"
            sku_val = str(sales_sheet.cell(r, s_sku).value or '').strip() if s_sku else ''
            if not sku_val: sku_val, _ = extract_sku_smart(msku, sku_set)
                
            sales_dict[key] = {
                '店铺': store, 'msku': msku,
                '销量': get_numeric_value(sales_sheet.cell(r, s_subtotal)),
                'SKU': sku_val,
                '品名': sales_sheet.cell(r, s_name).value if s_name else ''
            }

    # === 第3 & 4步：深圳仓 & 采购 ===
    sz_dict, po_dict = {}, {}
    if sz_stock_sheet_name:
        sz_sheet = wb[sz_stock_sheet_name]
        sz_sku = find_exact_col(sz_sheet, 'SKU') or 1
        sz_qty = find_exact_col(sz_sheet, '实际可用') or find_exact_col(sz_sheet, '可用') or 10
        empty_count = 0
        for r in range(2, sz_sheet.max_row + 1):
            sku = str(sz_sheet.cell(r, sz_sku).value or '').strip()
            if not sku:
                empty_count += 1
                if empty_count > 50: break
                continue
            empty_count = 0
            sz_dict[sku] = sz_dict.get(sku, 0) + get_numeric_value(sz_sheet.cell(r, sz_qty))

    if po_sheet_name:
        po_sheet = wb[po_sheet_name]
        po_sku = find_exact_col(po_sheet, 'SKU') or 7
        po_qty = find_exact_col(po_sheet, '未入库量') or find_exact_col(po_sheet, '未入库') or 19
        empty_count = 0
        for r in range(2, po_sheet.max_row + 1):
            sku = str(po_sheet.cell(r, po_sku).value or '').strip()
            if not sku:
                empty_count += 1
                if empty_count > 50: break
                continue
            empty_count = 0
            po_dict[sku] = po_dict.get(sku, 0) + get_numeric_value(po_sheet.cell(r, po_qty))

    # === 第5步：更新主表 ===
    i_store = find_exact_col(inv_sheet, '店铺') or 1
    i_msku = find_exact_col(inv_sheet, 'msku') or 2
    i_store_msku = find_exact_col(inv_sheet, '店铺&MSKU')
    i_gtin = find_exact_col(inv_sheet, 'GTIN码')
    i_name = find_exact_col(inv_sheet, '品名')
    i_sku = find_exact_col(inv_sheet, 'sku')
    i_status = find_exact_col(inv_sheet, '商品状态')
    
    i_avail = find_exact_col(inv_sheet, 'WFS可售(新)(数量)')
    i_unable = find_exact_col(inv_sheet, '无法入库(数量)')
    i_transit = find_exact_col(inv_sheet, '标发在途(数量)')
    i_sz = find_exact_col(inv_sheet, '深圳仓库存') or find_exact_col(inv_sheet, '深圳仓')
    i_po = find_exact_col(inv_sheet, '采购订单在途') or find_exact_col(inv_sheet, '采购')
    i_total = find_exact_col(inv_sheet, '总库存')
    i_turnover = find_exact_col(inv_sheet, '总周转天数（不含采购订单）') or find_exact_col(inv_sheet, '总周转')
    
    i_sales = find_exact_col(inv_sheet, sales_sheet_name)
    if not i_sales:
        i_sales = inv_sheet.max_column + 1
        inv_sheet.cell(2, i_sales, sales_sheet_name)

    # 5.1 更新历史行
    for r in range(3, original_max_row + 1):
        s = str(inv_sheet.cell(r, i_store).value or '').strip()
        m = str(inv_sheet.cell(r, i_msku).value or '').strip()
        if not s and not m: continue
        key = f"{s}{m}"
        
        curr_sku = str(inv_sheet.cell(r, i_sku).value or '').strip() if i_sku else ''

        if key in wfs_dict:
            d = wfs_dict[key]
            if i_gtin: inv_sheet.cell(r, i_gtin, d['GTIN码'])
            if i_name and d['品名']: inv_sheet.cell(r, i_name, d['品名'])
            if i_sku and d['sku']: 
                inv_sheet.cell(r, i_sku, d['sku'])
                curr_sku = d['sku']
            if i_status: inv_sheet.cell(r, i_status, d['商品状态'])
            if i_avail: inv_sheet.cell(r, i_avail, d['WFS可售'])
            if i_unable: inv_sheet.cell(r, i_unable, d['无法入库'])
            if i_transit: inv_sheet.cell(r, i_transit, d['标发在途'])
            
        if key in sales_dict and i_sales:
            inv_sheet.cell(r, i_sales, sales_dict[key]['销量'])
            
        if curr_sku:
            if i_sz and curr_sku in sz_dict: inv_sheet.cell(r, i_sz, sz_dict[curr_sku])
            if i_po and curr_sku in po_dict: inv_sheet.cell(r, i_po, po_dict[curr_sku])

    # 5.2 准备新增行
    all_keys = set(wfs_dict.keys()) | set(sales_dict.keys())
    new_rows = []
    
    for key in all_keys:
        if key in existing_keys: continue
        
        d = {'店铺': '', 'msku': '', '店铺&MSKU': key, 'sku': '', '品名': '',
             'WFS可售': 0, '无法入库': 0, '标发在途': 0, '销量': 0}
        
        if key in wfs_dict:
            d.update(wfs_dict[key])
            d['店铺'] = wfs_dict[key]['仓库']
        
        if key in sales_dict:
            d['销量'] = sales_dict[key]['销量']
            if not d['店铺']: d['店铺'] = sales_dict[key]['店铺']
            if not d['msku']: d['msku'] = sales_dict[key]['msku']
            if not d['sku']: d['sku'] = sales_dict[key]['SKU']
            if not d['品名']: d['品名'] = sales_dict[key]['品名']

        # 如果没有SKU，尝试通过MSKU提取
        if not d['sku']:
            ext_sku, _ = extract_sku_smart(d['msku'], sku_set)
            if ext_sku: d['sku'] = ext_sku
            
        # 补充品名
        if not d['品名'] and d['sku'] in sku_to_name:
            d['品名'] = sku_to_name[d['sku']]

        new_rows.append(d)

    # 写入新增行
    curr_row = original_max_row + 1
    for d in new_rows:
        if i_store: inv_sheet.cell(curr_row, i_store, d['店铺'])
        if i_msku: inv_sheet.cell(curr_row, i_msku, d['msku'])
        if i_store_msku: inv_sheet.cell(curr_row, i_store_msku, d['店铺&MSKU'])
        if i_gtin and 'GTIN码' in d: inv_sheet.cell(curr_row, i_gtin, d['GTIN码'])
        if i_sku: inv_sheet.cell(curr_row, i_sku, d['sku'])
        if i_name: inv_sheet.cell(curr_row, i_name, d['品名'])
        if i_status and '商品状态' in d: inv_sheet.cell(curr_row, i_status, d['商品状态'])
        
        if i_avail: inv_sheet.cell(curr_row, i_avail, d['WFS可售'])
        if i_unable: inv_sheet.cell(curr_row, i_unable, d['无法入库'])
        if i_transit: inv_sheet.cell(curr_row, i_transit, d['标发在途'])
        if i_sales: inv_sheet.cell(curr_row, i_sales, d['销量'])
        
        sku = d['sku']
        if sku:
            if i_sz and sku in sz_dict: inv_sheet.cell(curr_row, i_sz, sz_dict[sku])
            if i_po and sku in po_dict: inv_sheet.cell(curr_row, i_po, po_dict[sku])
            
        curr_row += 1

    # === 第6步：计算与严格清理 ===
    for r in range(3, curr_row):
        v_wfs = get_numeric_value(inv_sheet.cell(r, i_avail)) if i_avail else 0
        v_unable = get_numeric_value(inv_sheet.cell(r, i_unable)) if i_unable else 0
        v_transit = get_numeric_value(inv_sheet.cell(r, i_transit)) if i_transit else 0
        v_sz = get_numeric_value(inv_sheet.cell(r, i_sz)) if i_sz else 0
        v_sales = get_numeric_value(inv_sheet.cell(r, i_sales)) if i_sales else 0
        
        if i_total: inv_sheet.cell(r, i_total, v_wfs + v_unable + v_transit + v_sz)
        
        if i_turnover:
            if v_sales > 0:
                inv_sheet.cell(r, i_turnover, round((v_wfs + v_transit + v_sz) / v_sales * 30, 2))
            else:
                inv_sheet.cell(r, i_turnover, "")

    # 清理：仅针对新增区域，仅看4个核心数值
    rows_to_del = []
    for r in range(original_max_row + 1, curr_row):
        v_wfs = get_numeric_value(inv_sheet.cell(r, i_avail)) if i_avail else 0
        v_unable = get_numeric_value(inv_sheet.cell(r, i_unable)) if i_unable else 0
        v_transit = get_numeric_value(inv_sheet.cell(r, i_transit)) if i_transit else 0
        v_sales = get_numeric_value(inv_sheet.cell(r, i_sales)) if i_sales else 0
        
        # 核心更新：只需判断这4个数值是否均为0
        if v_wfs == 0 and v_unable == 0 and v_transit == 0 and v_sales == 0:
            rows_to_del.append(r)

    # 从下往上删，防止错位
    for r in sorted(rows_to_del, reverse=True):
        inv_sheet.delete_rows(r, 1)

    st.success(f"✅ 处理完毕！成功新增 {len(new_rows) - len(rows_to_del)} 条有效记录。")
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ==========================================
# Streamlit UI
# ==========================================
st.set_page_config(page_title="沃尔玛库存工具 v14", layout="wide")
st.title("🛒 沃尔玛库存更新工具 (V14 精确过滤版)")
st.markdown("""
**本次更新重点 (V14)：**
1. **防死循环且不漏数据**：加入“连续 50 行空探测器”，兼容数据内部缺漏，完美解决转圈卡死。
2. **全新精准过滤机制**：仅考核 `WFS可售`、`无法入库`、`标发在途` 和 `销量` 这 4 项。全为 0 即删除该新增行，不再受深圳仓及采购数量干扰，大幅减少冗余行。
""")

c1, c2 = st.columns(2)
with c1:
    f_inv = st.file_uploader("上传库存明细表 (必选)", type=['xlsx'], key="inv")
with c2:
    f_prod = st.file_uploader("上传产品资料表 (推荐)", type=['xlsx'], key="prod")

if f_inv and st.button("🚀 极速处理"):
    with st.spinner("正在光速为您计算合并中，请稍候..."):
        try:
            data = process_inventory(f_inv, f_prod)
            if data:
                st.balloons()
                st.download_button("📥 下载最终版库存表", data, f"Updated_{f_inv.name}")
        except Exception as e:
            st.error(f"处理发生异常，请检查表格格式: {e}")
