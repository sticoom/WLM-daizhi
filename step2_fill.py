import openpyxl
import io
import re
import warnings

warnings.filterwarnings('ignore')

def clean_header(val):
    if not val: return ""
    s = str(val).strip().lower()
    s = s.replace('（', '(').replace('）', ')')
    s = re.sub(r'\s+', '', s) 
    return s

def find_col(sheet, exact_names, excludes=None):
    excludes = excludes or []
    # 1. 优先精确匹配
    for col in range(1, 150):
        for r in [1, 2]:
            val = sheet.cell(row=r, column=col).value
            if not val: continue
            cleaned = clean_header(val)
            for exact in exact_names:
                if clean_header(exact) == cleaned:
                    return col
                    
    # 2. 安全模糊匹配
    for col in range(1, 150):
        for r in [1, 2]:
            val = sheet.cell(row=r, column=col).value
            if not val: continue
            cleaned = clean_header(val)
            for exact in exact_names:
                target = clean_header(exact)
                if target in cleaned:
                    if any(clean_header(ex) in cleaned for ex in excludes):
                        continue
                    if target == 'sku' and 'msku' in cleaned: continue
                    if target == 'msku' and '店铺' in cleaned: continue
                    if target == '店铺' and 'msku' in cleaned: continue
                    return col
    return None

def get_real_max_col(sheet, row_idx=2):
    """寻找真实的最后一列，防止幽灵列导致新加的列跑到16384列外去"""
    max_c = 1
    for c in range(150, 0, -1):
        val = sheet.cell(row=row_idx, column=c).value
        if val is not None and str(val).strip() != "":
            max_c = c
            break
    return max_c

def get_numeric_value(cell):
    if cell is None or cell.value is None: return 0
    val = str(cell.value).strip()
    if val == '' or val.lower() in ('nan', '#n/a', 'none'): return 0
    try:
        if val.startswith('='): return 0
        return float(val.replace(',', ''))
    except: return 0

def load_product_reference(product_file_obj):
    sku_set, sku_to_name = set(), {}
    if product_file_obj is None: return sku_set, sku_to_name
    try:
        wb = openpyxl.load_workbook(product_file_obj, read_only=True)
        ws = wb[wb.sheetnames[0]]
        sku_col, name_col = None, None
        for col in range(1, 20):
            val = str(ws.cell(1, col).value or '').lower()
            if 'sku' in val: sku_col = col
            if val in ['品名', '名称', 'name', '品名英']: name_col = col

        if sku_col:
            empty_count = 0
            for r in range(2, ws.max_row + 1):
                sku_val = str(ws.cell(r, sku_col).value or '').strip()
                if not sku_val:
                    empty_count += 1
                    if empty_count > 50: break
                    continue
                empty_count = 0
                sku_set.add(sku_val)
                if name_col:
                    name_val = str(ws.cell(r, name_col).value or '').strip()
                    if name_val: sku_to_name[sku_val] = name_val
        wb.close()
        return sku_set, sku_to_name
    except: return set(), {}

def extract_sku_smart(msku, sku_set):
    if not msku: return ''
    parts = [p.strip() for p in msku.split('-') if p.strip()]
    if not parts: return ''
    
    if not sku_set:
        return parts[1] if len(parts) >= 2 else parts[0]
        
    for p in parts:
        if p in sku_set: return p
        
    cleaned = [p.replace('"', '').replace("'", '').replace(' ', '') for p in parts]
    for p in cleaned:
        if p in sku_set: return p
        
    for p in cleaned:
        if len(p) >= 4 and re.search(r'\d', p) and re.search(r'[a-zA-Z]', p):
            for sku in sku_set:
                if p in sku or sku in p:
                    if len(sku) > 0 and len(p)/len(sku) >= 0.6: return sku
                    
    candidates = [p for p in cleaned if re.search(r'\d', p) and re.search(r'[a-zA-Z]', p)]
    if candidates:
        return max(candidates, key=len)
        
    return parts[1] if len(parts) >= 2 else parts[0]

def step2_fill_and_calculate(intermediate_file, product_file_obj):
    sku_set, sku_to_name = load_product_reference(product_file_obj)
    wb = openpyxl.load_workbook(intermediate_file)
    sheets = wb.sheetnames
    
    def find_sheet(kws):
        for s in sheets:
            if any(k in s for k in kws): return s
        return None

    inv_sheet_name = sheets[1]
    wfs_sheet_name = find_sheet(['WFS库存', 'WFS'])
    sz_sheet_name = find_sheet(['深圳仓', '深圳'])
    po_sheet_name = find_sheet(['采购订单', '采购', '在途'])
    sales_sheet_name = sheets[4] if len(sheets) > 4 else None

    # --- 获取 WFS 全量数据 ---
    wfs_full = {}
    if wfs_sheet_name:
        w_sh = wb[wfs_sheet_name]
        c_wh = find_col(w_sh, ['仓库', 'warehouse']) or 1
        c_msku = find_col(w_sh, ['msku']) or 2
        
        c_pid = find_col(w_sh, ['平台商品ID', '商品ID'], excludes=['gtin'])
        c_gtin = find_col(w_sh, ['GTIN码', 'gtin'])
        c_name = find_col(w_sh, ['品名'])
        c_sku = find_col(w_sh, ['sku'], excludes=['msku'])
        c_status = find_col(w_sh, ['商品状态'])
        
        c_avail = find_col(w_sh, ['WFS可售(新)(数量)'])
        c_unable = find_col(w_sh, ['无法入库(数量)'])
        c_transit = find_col(w_sh, ['标发在途(数量)'])
        c_a3 = find_col(w_sh, ['3个月内库龄(数量)'])
        c_a6 = find_col(w_sh, ['3-6个月库龄(数量)'])
        c_a6_9 = find_col(w_sh, ['6-9个月库龄(数量)'])
        c_a9_12 = find_col(w_sh, ['9-12个月库龄(数量)'])
        c_a6_plus = find_col(w_sh, ['6个月以上库龄(数量)'])
        c_a12p = find_col(w_sh, ['12个月以上库龄(数量)'])

        for r in range(2, w_sh.max_row + 1):
            wh = str(w_sh.cell(r, c_wh).value or '').strip()
            msku = str(w_sh.cell(r, c_msku).value or '').strip()
            if wh and msku:
                v_a6_9 = get_numeric_value(w_sh.cell(r, c_a6_9)) if c_a6_9 else 0
                v_a9_12 = get_numeric_value(w_sh.cell(r, c_a9_12)) if c_a9_12 else 0
                v_12p = get_numeric_value(w_sh.cell(r, c_a12p)) if c_a12p else 0
                v_6plus = get_numeric_value(w_sh.cell(r, c_a6_plus)) if c_a6_plus else (v_a6_9 + v_a9_12 + v_12p)
                
                wfs_full[f"{wh}{msku}"] = {
                    '平台商品ID': str(w_sh.cell(r, c_pid).value or '') if c_pid else '',
                    'GTIN': str(w_sh.cell(r, c_gtin).value or '') if c_gtin else '',
                    '品名': str(w_sh.cell(r, c_name).value or '').strip() if c_name else '',
                    'SKU': str(w_sh.cell(r, c_sku).value or '').strip() if c_sku else '',
                    '状态': w_sh.cell(r, c_status).value if c_status else '',
                    '可售': get_numeric_value(w_sh.cell(r, c_avail)) if c_avail else 0,
                    '无法': get_numeric_value(w_sh.cell(r, c_unable)) if c_unable else 0,
                    '在途': get_numeric_value(w_sh.cell(r, c_transit)) if c_transit else 0,
                    '库龄3': get_numeric_value(w_sh.cell(r, c_a3)) if c_a3 else 0,
                    '库龄6': get_numeric_value(w_sh.cell(r, c_a6)) if c_a6 else 0,
                    '库龄12': v_6plus, 
                    '库龄超12': v_12p,
                }

    # --- 获取销量数据 ---
    sales_full = {}
    if sales_sheet_name:
        s_sh = wb[sales_sheet_name]
        s_msku = find_col(s_sh, ['MSKU']) or 4
        s_store = find_col(s_sh, ['店铺']) or 3
        s_subtotal = find_col(s_sh, ['小计']) or 13
        for r in range(2, s_sh.max_row + 1):
            msku = str(s_sh.cell(r, s_msku).value or '').strip()
            store = str(s_sh.cell(r, s_store).value or '').strip()
            if not store and '-' in msku: store = msku.split('-')[0]
            if msku:
                sales_full[f"{store}{msku}"] = get_numeric_value(s_sh.cell(r, s_subtotal))

    # --- 获取深圳与采购数据 ---
    sz_full, po_full = {}, {}
    if sz_sheet_name:
        sz_sh = wb[sz_sheet_name]
        sz_sku = find_col(sz_sh, ['SKU'], excludes=['msku']) or 1
        sz_qty = find_col(sz_sh, ['可用库存', '实际可用', '可用']) or 8
        sz_wh = find_col(sz_sh, ['仓库名称', '仓库']) or 4 
        
        for r in range(2, sz_sh.max_row + 1):
            sku = str(sz_sh.cell(r, sz_sku).value or '').strip()
            wh_name = str(sz_sh.cell(r, sz_wh).value or '').strip()
            if sku and '沃尔玛深圳仓' in wh_name: 
                sz_full[sku] = sz_full.get(sku, 0) + get_numeric_value(sz_sh.cell(r, sz_qty))

    if po_sheet_name:
        po_sh = wb[po_sheet_name]
        po_sku = find_col(po_sh, ['SKU'], excludes=['msku']) or 7
        po_qty = find_col(po_sh, ['未入库量', '未入库']) or 19
        po_req = find_col(po_sh, ['需求人']) or 28 
        
        for r in range(2, po_sh.max_row + 1):
            sku = str(po_sh.cell(r, po_sku).value or '').strip()
            requester = str(po_sh.cell(r, po_req).value or '').strip()
            if sku and '陈丹丹' in requester: 
                po_full[sku] = po_full.get(sku, 0) + get_numeric_value(po_sh.cell(r, po_qty))

    # --- 注入主表与计算 ---
    inv_sh = wb[inv_sheet_name]
    
    i_store = find_col(inv_sh, ['店铺'], excludes=['msku']) or 1
    i_msku = find_col(inv_sh, ['msku'], excludes=['店铺']) or 2
    i_sku = find_col(inv_sh, ['sku'], excludes=['msku'])
    
    i_pid = find_col(inv_sh, ['平台商品ID', '商品ID'], excludes=['gtin'])
    i_name = find_col(inv_sh, ['品名'])
    i_gtin = find_col(inv_sh, ['GTIN码'])
    i_status = find_col(inv_sh, ['商品状态'])
    
    i_avail = find_col(inv_sh, ['WFS可售(新)(数量)'])
    i_unable = find_col(inv_sh, ['无法入库(数量)'])
    i_transit = find_col(inv_sh, ['标发在途(数量)'])
    
    i_a3 = find_col(inv_sh, ['3个月内库龄(数量)'])
    i_a6 = find_col(inv_sh, ['3-6个月库龄(数量)'])
    i_a12 = find_col(inv_sh, ['6个月以上库龄(数量)'])
    i_a12p = find_col(inv_sh, ['12个月以上库龄(数量)'])
    
    i_sz = find_col(inv_sh, ['深圳仓库存', '深圳仓'])
    i_po = find_col(inv_sh, ['采购订单在途', '采购'])
    
    i_total = find_col(inv_sh, ['总库存（不含采购订单）', '总库存'])
    i_turn_wfs = find_col(inv_sh, ['WFS在库周转'])
    i_turn_transit = find_col(inv_sh, ['WFS在途+在库周转'])
    i_turn_total = find_col(inv_sh, ['总周转天数（不含采购订单）', '总周转'])
    
    # 【重点修复 1】：防幽灵列寻找真实最后列，并在表头拼接新增的“销量明细”列
    i_sales = find_col(inv_sh, [sales_sheet_name])
    if not i_sales:
        real_max_col = get_real_max_col(inv_sh, 2)
        i_sales = real_max_col + 1
        inv_sh.cell(2, i_sales, sales_sheet_name)

    empty_count = 0
    for r in range(3, inv_sh.max_row + 1):
        store = str(inv_sh.cell(r, i_store).value or '').strip()
        msku = str(inv_sh.cell(r, i_msku).value or '').strip()
        
        if not store and not msku:
            empty_count += 1
            if empty_count > 50: break
            continue
        empty_count = 0
        
        key = f"{store}{msku}"
        
        # 补全 SKU 和 品名
        curr_sku = str(inv_sh.cell(r, i_sku).value or '').strip() if i_sku else ''
        if not curr_sku and msku:
            curr_sku = extract_sku_smart(msku, sku_set)
            if not curr_sku and key in wfs_full and wfs_full[key]['SKU']:
                curr_sku = wfs_full[key]['SKU']
            if i_sku and curr_sku: 
                inv_sh.cell(r, i_sku, curr_sku)
        
        curr_name = str(inv_sh.cell(r, i_name).value or '').strip() if i_name else ''
        if not curr_name:
            if key in wfs_full and wfs_full[key]['品名']:
                curr_name = wfs_full[key]['品名']
            elif curr_sku in sku_to_name:
                curr_name = sku_to_name[curr_sku]
            if i_name and curr_name:
                inv_sh.cell(r, i_name, curr_name)
                
        # 更新该行数据
        if key in wfs_full:
            d = wfs_full[key]
            if i_pid and d['平台商品ID']: inv_sh.cell(r, i_pid, d['平台商品ID'])
            if i_gtin and d['GTIN']: inv_sh.cell(r, i_gtin, d['GTIN'])
            if i_status and d['状态']: inv_sh.cell(r, i_status, d['状态'])
            
            if i_avail: inv_sh.cell(r, i_avail, d['可售'])
            if i_unable: inv_sh.cell(r, i_unable, d['无法'])
            if i_transit: inv_sh.cell(r, i_transit, d['在途'])
            if i_a3: inv_sh.cell(r, i_a3, d['库龄3'])
            if i_a6: inv_sh.cell(r, i_a6, d['库龄6'])
            if i_a12: inv_sh.cell(r, i_a12, d['库龄12'])
            if i_a12p: inv_sh.cell(r, i_a12p, d['库龄超12'])
            
        if key in sales_full:
            if i_sales: inv_sh.cell(r, i_sales, sales_full[key])
            
        if curr_sku:
            if curr_sku in sz_full:
                if i_sz: inv_sh.cell(r, i_sz, sz_full[curr_sku])
            if curr_sku in po_full:
                if i_po: inv_sh.cell(r, i_po, po_full[curr_sku])
                
        # 【重点修复 2】：先从单元格中读取当前最新值，然后再进行严格的计算！
        v_wfs = get_numeric_value(inv_sh.cell(r, i_avail)) if i_avail else 0
        v_unable = get_numeric_value(inv_sh.cell(r, i_unable)) if i_unable else 0
        v_transit = get_numeric_value(inv_sh.cell(r, i_transit)) if i_transit else 0
        v_sz = get_numeric_value(inv_sh.cell(r, i_sz)) if i_sz else 0
        v_sales = get_numeric_value(inv_sh.cell(r, i_sales)) if i_sales else 0
                
        # 1. 计算总库存（不含采购订单） = WFS可售(新) + 无法入库 + 标发在途 + 深圳仓
        if i_total:
            inv_sh.cell(r, i_total, v_wfs + v_unable + v_transit + v_sz)
            
        # 2. 严格按 Skill 公式计算周转率
        if v_sales > 0:
            if i_turn_wfs: 
                inv_sh.cell(r, i_turn_wfs, round(v_wfs / v_sales * 30, 2))
            if i_turn_transit: 
                inv_sh.cell(r, i_turn_transit, round((v_wfs + v_transit) / v_sales * 30, 2))
            if i_turn_total: 
                inv_sh.cell(r, i_turn_total, round((v_wfs + v_transit + v_sz) / v_sales * 30, 2))
        else:
            # 销量为0时返回空值
            if i_turn_wfs: inv_sh.cell(r, i_turn_wfs, "")
            if i_turn_transit: inv_sh.cell(r, i_turn_transit, "")
            if i_turn_total: inv_sh.cell(r, i_turn_total, "")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
